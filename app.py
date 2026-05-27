import os, io, math, random, re
from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pdfplumber
import pandas as pd

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024

# ── Constants ─────────────────────────────────────────────────────────────────
PTS10  = [50, 25, 20, 17, 13, 10]
COST_U = [1470, 735, 588, 499.8, 382.2, 294]


# ── PDF parser (sin API key) ──────────────────────────────────────────────────
def parse_jca_pdf(file_obj):
    with pdfplumber.open(file_obj) as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)

    r = {}

    m = re.search(r'(NF\d{6,})', text)
    r['folio'] = m.group(1) if m else ''
    if not r['folio']:
        m = re.search(r'N+F+(\d{6,})', text)
        if m: r['folio'] = 'NF' + m.group(1)

    m = re.search(r'(\d{2}/[A-Z]{3}/\d{4})', text)
    r['fecha'] = m.group(1) if m else ''

    m = re.search(r'Usa\.\s+([\d,]+\.?\d*)\s+\$\s+294', text)
    r['points'] = round(float(m.group(1).replace(',', '')), 2) if m else 0.0

    m = re.search(r'\$\s*([\d,]+\.\d{2})\s*\.\s*Usa\.', text)
    r['subtotal'] = float(m.group(1).replace(',', '')) if m else round(r['points'] * 294, 2)

    m = re.search(r'TASA 16 %\s+\$?([\d,]+\.\d+)', text)
    r['iva'] = float(m.group(1).replace(',', '')) if m else round(r['subtotal'] * 0.16, 2)

    expected = r['subtotal'] * 1.16
    nums = [float(t.replace(',', '')) for t in re.findall(r'\$\s*([\d,]+\.\d{2})', text)]
    r['total'] = min(nums, key=lambda x: abs(x - expected), default=round(expected, 2)) if nums else round(expected, 2)

    m = re.search(r'R\.F\.C\.\s+(JCA\w+)', text)
    r['rfcEmisor'] = m.group(1) if m else 'JCA100604EF4'
    m = re.search(r'R\.F\.C\.\s+(EPA\w+)', text)
    r['rfcReceptor'] = m.group(1) if m else 'EPA191009JJ0'

    return r


# ── Stock parser ──────────────────────────────────────────────────────────────
def parse_stock(file_obj):
    df = pd.read_excel(file_obj, header=0)
    stock = {f'G{i}': 0 for i in range(1, 8)}
    for _, row in df.iterrows():
        prod = str(row.iloc[1]) if len(row) > 1 else ''
        qty  = row.iloc[2]     if len(row) > 2 else 0
        m = re.search(r'\[G(\d)', prod)
        if m:
            g = f'G{m.group(1)}'
            if g in stock:
                try: stock[g] += int(qty)
                except: pass
    return stock


# ── Combination generator ─────────────────────────────────────────────────────
def cv_score(combo):
    m = sum(combo) / len(combo)
    if not m: return 999
    return math.sqrt(sum((x - m)**2 for x in combo) / len(combo)) / m


def generate_combinations(target10, max_q, n=50, seed=42):
    random.seed(seed)
    results = []; seen = set()
    min_r = lambda f: sum(PTS10[f:])

    for _ in range(4_000_000):
        if len(results) >= n * 8: break
        mx67 = target10 // 50
        if mx67 < 1: break
        g67 = random.randint(1, min(max_q[0], mx67))
        r1  = target10 - g67 * 50
        if r1 < min_r(1): continue
        g5  = random.randint(1, min(max_q[1], r1 // 25))
        r2  = r1 - g5 * 25
        if r2 < min_r(2): continue
        g4  = random.randint(1, min(max_q[2], r2 // 20))
        r3  = r2 - g4 * 20
        if r3 < min_r(3): continue
        placed = False
        for dg3 in range(0, -9, -1):
            g3 = min(max_q[3], r3 // 17) + dg3
            if g3 < 1: break
            r4 = r3 - g3 * 17
            if r4 < PTS10[4] + PTS10[5]: continue
            for dg2 in range(0, -9, -1):
                g2 = min(max_q[4], r4 // 13) + dg2
                if g2 < 1: break
                r5 = r4 - g2 * 13
                if r5 >= 10 and r5 % 10 == 0:
                    g1 = r5 // 10
                    if 1 <= g1 <= max_q[5]:
                        k = (g67, g5, g4, g3, g2, g1)
                        if k not in seen:
                            seen.add(k); results.append(k)
                        placed = True; break
            if placed: break

    results.sort(key=cv_score)
    final = []; seen2 = set()
    for c in results:
        if c not in seen2:
            seen2.add(c); final.append(list(c))
        if len(final) >= n: break
    return final


# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(combos, stock, info, max_q):
    FILL = lambda c: PatternFill("solid", start_color=c)
    THIN = Side(style='thin', color="BBBBBB")
    BRD  = lambda: Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    EP_DARK  = "0D1F35"
    EP_MID   = "1E3A5F"
    EP_GREEN = "00C864"
    GRP_COLORS = ["4AADAD","D4A017","D96B64","5CA869","8E6BA0","D4834A"]

    def hc(ws, r, c, v, bg=EP_DARK, sz=10, wt=False):
        x = ws.cell(row=r, column=c, value=v)
        x.font = Font(bold=True, name="Calibri", size=sz, color="FFFFFF")
        x.fill = FILL(bg); x.border = BRD()
        x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wt)
        return x

    def dc(ws, r, c, v, bg="FFFFFF", bold=False, fmt=None, color="000000"):
        x = ws.cell(row=r, column=c, value=v)
        x.font = Font(bold=bold, name="Calibri", size=10, color=color)
        x.fill = FILL(bg); x.border = BRD()
        x.alignment = Alignment(horizontal="center", vertical="center")
        if fmt: x.number_format = fmt
        return x

    wb = Workbook()

    # Sheet 1: Info
    ws1 = wb.active; ws1.title = "Información"
    for i, w in enumerate([22, 22, 18, 16], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.merge_cells("A1:D1")
    ws1["A1"] = "ENERGY PARTS — GROUPS OPTIMIZER"
    ws1["A1"].font = Font(bold=True, name="Calibri", size=16, color="FFFFFF")
    ws1["A1"].fill = FILL(EP_DARK)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:D2")
    ws1["A2"] = f"Nota: {info.get('folio','')}  ·  {info.get('points',0)} pts  ·  {len(combos)} combinaciones"
    ws1["A2"].font = Font(italic=True, name="Calibri", size=10, color=EP_GREEN)
    ws1["A2"].fill = FILL(EP_MID)
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")

    campos = [
        ("Nota de Crédito", info.get("folio","")),
        ("Fecha", info.get("fecha","")),
        ("RFC Emisor", info.get("rfcEmisor","")),
        ("RFC Receptor", info.get("rfcReceptor","")),
        ("Subtotal sin IVA", info.get("subtotal", 0)),
        ("IVA 16%", info.get("iva", 0)),
        ("Total MXN", info.get("total", 0)),
        ("Puntos (÷$294)", info.get("points", 0)),
    ]
    for i, (lbl, val) in enumerate(campos, 4):
        c1 = ws1.cell(i, 1, lbl)
        c1.font = Font(bold=True, name="Calibri", size=10, color="595959")
        c1.fill = FILL("F0F4F8"); c1.border = BRD()
        c2 = ws1.cell(i, 2, val)
        c2.font = Font(bold="pts" in str(lbl).lower(), name="Calibri", size=10,
                       color=EP_MID if "pts" in str(lbl).lower() else "000000")
        c2.fill = FILL("FFFDE7" if "pts" in str(lbl).lower() else "FFFFFF")
        c2.border = BRD()
        ws1.row_dimensions[i].height = 16

    r = 13
    ws1.merge_cells(f"A{r}:D{r}")
    ws1.cell(r, 1, "EXISTENCIAS").font = Font(bold=True, name="Calibri", size=11, color=EP_DARK)
    ws1.cell(r, 1).alignment = Alignment(horizontal="center")
    for c, h in enumerate(["Grupo","Stock","Pts/u","Costo/u ($)"], 1):
        hc(ws1, r+1, c, h, bg=EP_MID)
    for i, (g, p, co) in enumerate([("G1",1.0,294),("G2",1.3,382.2),("G3",1.7,499.8),
                                      ("G4",2.0,588),("G5",2.5,735),("G6",5.0,1470),("G7",5.0,1470)], r+2):
        bg = "F8F8F8" if i%2==0 else "FFFFFF"
        dc(ws1,i,1,g,bg=bg,bold=True,color=EP_DARK)
        dc(ws1,i,2,stock.get(g,0),bg=bg,bold=True)
        dc(ws1,i,3,p,bg=bg,fmt='0.0" pts"')
        dc(ws1,i,4,co,bg=bg,fmt='"$"#,##0.00')
    dc(ws1,r+9,1,"G6+G7 POOL",bg="D4EDDA",bold=True,color=EP_MID)
    dc(ws1,r+9,2,stock.get("G6",0)+stock.get("G7",0),bg="D4EDDA",bold=True)
    dc(ws1,r+9,3,5.0,bg="D4EDDA",fmt='0.0" pts"')
    dc(ws1,r+9,4,1470,bg="D4EDDA",fmt='"$"#,##0.00')

    # Sheet 2: Combinations
    ws2 = wb.create_sheet("Combinaciones")
    for i, w in enumerate([4,12,11,10,10,10,10,10,11,14,14,9], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.merge_cells("A1:L1")
    ws2["A1"] = f"ENERGY PARTS  ·  {info.get('folio','')}  ·  {info.get('points',0)} pts  ·  {len(combos)} combinaciones"
    ws2["A1"].font = Font(bold=True, name="Calibri", size=13, color="FFFFFF")
    ws2["A1"].fill = FILL(EP_DARK)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    ws2.merge_cells("A2:L2")
    stock_str = "  ·  ".join(f"{g}: {stock.get(g,0)}" for g in ["G1","G2","G3","G4","G5","G6","G7"])
    ws2["A2"] = f"STOCK  →  {stock_str}  ·  G6+G7: {stock.get('G6',0)+stock.get('G7',0)}"
    ws2["A2"].font = Font(italic=True, name="Calibri", size=9, color=EP_GREEN)
    ws2["A2"].fill = FILL(EP_MID)
    ws2["A2"].alignment = Alignment(horizontal="center", vertical="center")

    hdrs = ["#","G6/G7\n(5pts)","G5\n(2.5)","G4\n(2)","G3\n(1.7)","G2\n(1.3)","G1\n(1)",
            "Total\nGrupos","Total\nPuntos","Total\n$Costo","Uniformidad","Stock\n✓/✗"]
    for c, h in enumerate(hdrs, 1):
        hc(ws2, 3, c, h, bg=EP_DARK, wt=True)
    ws2.row_dimensions[3].height = 32

    hc(ws2, 4, 1, "MÁX →", bg=EP_MID, sz=9)
    for j, v in enumerate(max_q, 2):
        x = ws2.cell(4, j, v)
        x.font = Font(bold=True, name="Calibri", size=9, color="FFFFFF")
        x.fill = FILL(EP_MID); x.border = BRD()
        x.alignment = Alignment(horizontal="center", vertical="center")
    for j in range(8, 13):
        ws2.cell(4, j).fill = FILL(EP_MID); ws2.cell(4, j).border = BRD()
    ws2.row_dimensions[4].height = 14

    for i, combo in enumerate(combos, 1):
        r = 4 + i
        g67, g5, g4, g3, g2, g1 = combo
        vals = [g67, g5, g4, g3, g2, g1]
        tu  = sum(vals)
        tp  = (g67*50+g5*25+g4*20+g3*17+g2*13+g1*10)/10
        tc  = g67*1470+g5*735+g4*588+g3*499.8+g2*382.2+g1*294
        cv  = cv_score(vals)
        badge = "★ MUY UNIFORME" if cv<0.50 else "✓ UNIFORME" if cv<0.65 else "VÁLIDA"
        bcol  = "00A550" if cv<0.50 else "1E5FAF" if cv<0.65 else "D97B00"
        ok    = all(vals[j] <= max_q[j] for j in range(6))
        row_bg = "F5FAF7" if i%2==0 else "FFFFFF"

        dc(ws2, r, 1, i, bg=row_bg)
        for j, (v, gc) in enumerate(zip(vals, GRP_COLORS), 2):
            x = ws2.cell(r, j, v)
            x.font = Font(bold=True, name="Calibri", size=10,
                          color="C00000" if v > max_q[j-2] else "000000")
            x.fill = FILL(gc+"22" if v<=max_q[j-2] else "FFE0E0")
            x.border = BRD(); x.alignment = Alignment(horizontal="center", vertical="center")
        dc(ws2, r, 8, tu, bg=row_bg, bold=True)
        dc(ws2, r, 9, tp, bg=row_bg, bold=True, fmt='#,##0.0" pts"')
        dc(ws2, r, 10, tc, bg=row_bg, bold=True, fmt='"$"#,##0.00')
        bx = ws2.cell(r, 11, badge)
        bx.font = Font(bold=True, name="Calibri", size=9, color=bcol)
        bx.fill = FILL("E8F5EE" if cv<0.50 else "EEF4FF" if cv<0.65 else "FFF8EC")
        bx.border = BRD(); bx.alignment = Alignment(horizontal="center", vertical="center")
        sx = ws2.cell(r, 12, "✓" if ok else "✗")
        sx.font = Font(bold=True, name="Calibri", size=12, color="00A550" if ok else "C00000")
        sx.fill = FILL(row_bg); sx.border = BRD()
        sx.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r].height = 15

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/extract-pdf", methods=["POST"])
def extract_pdf():
    if "pdf" not in request.files:
        return jsonify({"error": "No se recibió PDF"}), 400
    try:
        info = parse_jca_pdf(request.files["pdf"])
        return jsonify(info)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/parse-stock", methods=["POST"])
def parse_stock_route():
    if "stock" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400
    try:
        return jsonify(parse_stock(request.files["stock"]))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/generate", methods=["POST"])
def generate_route():
    data     = request.json
    stock    = data["stock"]
    info     = data["info"]
    n        = int(data.get("numCombos", 50))
    target10 = round(float(info["points"]) * 10)
    max_q    = [stock["G6"]+stock["G7"], stock["G5"], stock["G4"],
                stock["G3"], stock["G2"], stock["G1"]]
    combos   = generate_combinations(target10, max_q, n)
    return jsonify({"combos": combos, "count": len(combos)})


@app.route("/api/download-excel", methods=["POST"])
def download_excel():
    data   = request.json
    stock  = data["stock"]
    info   = data["info"]
    combos = data["combos"]
    max_q  = [stock["G6"]+stock["G7"], stock["G5"], stock["G4"],
              stock["G3"], stock["G2"], stock["G1"]]
    buf    = build_excel(combos, stock, info, max_q)
    fname  = f"{info.get('folio','combinaciones')}_grupos.xlsx"
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fname)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
